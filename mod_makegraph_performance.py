from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from openpyxl.chart.axis import ChartLines
import win32com.client
from collections import Counter
import numpy as np
import openpyxl as xl
import sys
from tkinter import messagebox



"""
測定データを引き切りのデータと背圧特性のデータに区別する．

第1引数：測定データのワークシート
第2引数；測定データが始まる行番号．（測定時のエクセルシートのフォーマットが変わらなければ18となる）
"""
def data_sort(measured_data_s, start_row):
    # 最終行を取得．この場合，数値が入力されている最終行ではなく，書式設定されているセルの最終行．
    maxrow = measured_data_s.max_row + 1

    # 流量の列(C列)を反対から進んで初めて空白セルじゃないところを最終行としている．
    for i in reversed(range(1, maxrow)):
        if measured_data_s.cell(i,3).value != None:
            end_row = i
            break

    # 引き切りのデータに該当する行番号を格納するリスト．
    list_rownum_pspq = []
    # 背圧特性のデータに該当する行番号を格納するリスト．
    list_rownum_backpressure = []
    for i in range(start_row, end_row + 1):
        # まず，SRGの列にデータがない行と流量の列が0，もしくは空白の行を除外．
        if (measured_data_s.cell(i,7).value != None) and (measured_data_s.cell(i,3).value != (None or 0)):
            # 排気口圧力の列が空白の行の行番号をリストに格納．
            if measured_data_s.cell(i,6).value == None:
                list_rownum_pspq.append(i)
            # 空白ではない行の行番号をリストに格納．
            else: list_rownum_backpressure.append(i)
    pspq = None
    backpressure = None

    if list_rownum_pspq != []:
        # 引き切りのデータの，[流量，SRG値，VAT値，流速]のリストを作成．
        list_pspq = []
        for i in list_rownum_pspq:
            list_pspq.append([measured_data_s.cell(i,3).value, measured_data_s.cell(i,7).value, measured_data_s.cell(i,11).value,measured_data_s.cell(i,12).value])
        pspq = True

    if list_rownum_backpressure != []:
        # ここから背圧特性のデータ整理
        # C列からsccmを取得
        list_sccm = []
        for i in list_rownum_backpressure:
            list_sccm.append(round(measured_data_s.cell(i,3).value))

        # list_sccmの中から種類を抜き出す．（辞書｛sccm：個数｝を作成し，key()で種類を抜き出す）
        list_sccm_var = Counter(list_sccm).keys()

        # 背圧特性評価に必要なデータを辞書形式で保存．
        # key = 流量, value = [排気口圧力，吸気口圧力]
        dict_backpressure = {}
        for i in sorted(list_sccm_var):
            dict_backpressure[i] = []
            for j in list_rownum_backpressure:
                if round(measured_data_s.cell(j,3).value) == i:
                    dict_backpressure[i].append([measured_data_s.cell(j,11).value, measured_data_s.cell(j,7).value])
            dict_backpressure[i] = np.array(dict_backpressure[i])
        backpressure = True

    if pspq == None and backpressure == None:
        messagebox.showerror('エラー！','グラフを作成するためのデータがありません!.')
        sys.exit()

    elif pspq == True and backpressure == None:
        return pspq, backpressure, list_pspq, None
    elif pspq == None and backpressure == True:
        return pspq, backpressure, None, dict_backpressure
    else:
        return pspq, backpressure, list_pspq, dict_backpressure




"""
PSPQ曲線用のデータ整理用のシートを作成するモジュール

第1引数：シート
第2引数：試験コンフィグ
第3引数：引き切りのデータが格納しているリスト
第4引数：SRGの単位
"""
def write_pspq_data(ws, test_config, list_pspq, dim_srg):
    list_word = [['','',test_config,'','','',''],['Gas throughput', '', 'Inlet pressure', '', 'Foreline pressure', '', 'Pumping speed'], ['sccm', '', 'Torr', 'Pa', 'Torr', 'Pa', 'L/s']]

    if dim_srg == 'Torr':
        coef_mat = [1,133.32]
    elif dim_srg == 'Pa':
        coef_mat = [1/133.32, 1]
    else:
        messagebox.showerror('SRGの単位が正しく記入されていません')

    for i in range(len(list_word)):
        for j in range(len(list_word[0])):
            ws.cell(i+1,j+1,value= list_word[i][j])

    for i in range(len(list_pspq)):
        ws.cell(i+4,1,value=list_pspq[i][0])
        ws.cell(i+4,3,value=list_pspq[i][1]*coef_mat[0])
        ws.cell(i+4,4,value=list_pspq[i][1]*coef_mat[1])
        ws.cell(i+4,5,value=list_pspq[i][2])
        ws.cell(i+4,6,value=list_pspq[i][2])
        ws.cell(i+4,7,value=list_pspq[i][3])



"""
グラフを作成するモジュール．

第1引数:シート
第2引数:試験コンフィグ．これが系列名となる．

"""
def make_pspq_curve(ws, test_config):
    # フォント設定．ややこしいので以下URL参照
    # https://hk29.hatenablog.jp/entry/2019/11/09/175122
    # https://syachiku-python.com/%E3%80%90%E5%AE%8C%E5%85%A8%E7%89%88%E3%80%91python-%E3%82%A8%E3%82%AF%E3%82%BB%E3%83%AB%E3%81%A7%E3%82%B0%E3%83%A9%E3%83%95%E3%82%92%E4%BD%9C%E3%82%8B-%E8%A7%A3%E8%AA%AC%E4%BB%98%E3%81%8D%E3%80%907/#toc10
    char_properties = CharacterProperties(latin=Font(typeface='Meiryo UI'), sz=11*100, b=False, solidFill="000000")
    paragraph_properties = ParagraphProperties(defRPr=char_properties)
    rich_text = RichText(p=[Paragraph(pPr=paragraph_properties, endParaRPr=char_properties)])

    # 散布図を作成
    ps = ScatterChart()
    # グラフのサイズを変更
    ps.height = 12
    ps.width = 20
    # タイトルはなし（バグかわからないがタイトルが消えない）
    ps.title = None
    # 軸ラベル設定
    ps.x_axis.title = 'Inlet pressure [Pa]'
    ps.y_axis.title = 'Pumping speed [L/s]'

    # column_sccm = list(ws.iter_rows(min_col=4, min_row=4, max_row=ws.max_row).values)
    # print(f'{column_sccm=}')

    # 参照するデータを設定
    x_values = Reference(ws, min_col=4, min_row=4, max_row=ws.max_row)
    y_values = Reference(ws, min_col=7, min_row=4, max_row=ws.max_row)
    series = Series(y_values, x_values, title= test_config)
    ps.series.append(series)

    # 軸ラベルや目盛のフォント変更
    ps.x_axis.txPr = rich_text
    ps.x_axis.title.tx.rich.p[0].r[0].rPr = char_properties
    ps.y_axis.txPr = rich_text
    ps.y_axis.title.tx.rich.p[0].r[0].rPr = char_properties

    # X軸を対数軸にする
    ps.x_axis.scaling.logBase = 10
    # 軸の位置は下端，左端に設定
    ps.x_axis.tickLblPos = "low"
    ps.y_axis.tickLblPos = "low"

    # 補助目盛を追加
    ps.x_axis.minorGridlines = ChartLines()

    # 凡例の位置とフォント
    ps.legend.legendPos = "tr"
    ps.legend.txPr = rich_text

    # ps曲線をシートに追加
    ws.add_chart(ps, 'B20')

    # 以下，PQ曲線作成．上記同様．
    pq = ScatterChart()
    x_values = Reference(ws, min_col=4, min_row=4, max_row=ws.max_row)
    y_values = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
    series = Series(y_values, x_values, title= test_config)
    pq.series.append(series)
    pq.height = 12
    pq.width = 20
    pq.title = None
    pq.x_axis.title = 'Inlet pressure [Pa]'
    pq.y_axis.title = 'Gas throughput [sccm]'
    pq.x_axis.txPr = rich_text
    pq.y_axis.txPr = rich_text
    pq.x_axis.title.tx.rich.p[0].r[0].rPr = char_properties
    pq.y_axis.title.tx.rich.p[0].r[0].rPr = char_properties
    pq.x_axis.scaling.logBase = 10
    pq.x_axis.tickLblPos = "low"
    pq.y_axis.tickLblPos = "low"
    pq.x_axis.minorGridlines = ChartLines()
    pq.legend.legendPos = "tr"
    pq.legend.txPr = rich_text

    ws.add_chart(pq, 'O20')


"""
背圧特性のデータ整理シートを作成するモジュール

第1引数：ワークシート
第2引数：試験コンフィグ
第3引数：背圧特性用のデータが格納している辞書
第4引数：SRGの単位
"""
def write_backpressure_data(ws, test_config, dict_backpressure, dim_srg):
    if dim_srg == 'Torr':
        coef_mat = [1,133.32]
    elif dim_srg == 'Pa':
        coef_mat == [1/133.32, 1]
    else:
        messagebox('エラー！','SRGの単位が選択されていません')

    sccm_list = list(dict_backpressure.keys())

    ws.cell(1,1,value= test_config)
    for i in range(len(sccm_list)):
        ws.cell(2,2*i+1,value=sccm_list[i])
        ws.cell(3,2*i+1, value='Outlet pressure\n[Torr]')
        ws.cell(3,2*i+2, value='Inlet pressure\n[Torr]')
        for j in range(len(dict_backpressure[sccm_list[i]])):
            ws.cell(j+4,2*i+1, value=dict_backpressure[sccm_list[i]][j,0])
            ws.cell(j+4,2*i+2, value=dict_backpressure[sccm_list[i]][j,1])


"""
背圧特性のグラフを作成するモジュール

第1引数：ワークシート
第2引数：試験コンフィグ
第3引数：背圧特性のデータを格納している辞書
"""
def make_backpressure_curve(ws, test_config, dict_backpressure):
    char_properties = CharacterProperties(latin=Font(typeface='Meiryo UI'), sz=11*100, b=False, solidFill="000000")
    paragraph_properties = ParagraphProperties(defRPr=char_properties)
    rich_text = RichText(p=[Paragraph(pPr=paragraph_properties, endParaRPr=char_properties)])


    # 散布図を作成
    chart = ScatterChart()
    # グラフのサイズを変更
    chart.height = 24
    chart.width = 16
    # タイトルはなし（バグかわからないがタイトルが消えないかも）
    chart.title = None
    # 軸ラベル設定
    chart.x_axis.title = 'Outlet pressure [Pa]'
    chart.y_axis.title = 'Inlet pressure [Pa]'

    # 参照するデータを設定
    sccm_list = list(dict_backpressure.keys())
    num_series = len(sccm_list)
    for i in range(num_series):
        x_values = Reference(ws, min_col=2*i+1, min_row=4, max_row=3 + len(dict_backpressure[sccm_list[i]]))
        y_values = Reference(ws, min_col=2*i+2, min_row=4, max_row=3 + len(dict_backpressure[sccm_list[i]]))
        series = Series(y_values, x_values, title=test_config + '-%dsccm' %sccm_list[i])
        chart.series.append(series)

    # 軸ラベルや目盛のフォント変更
    chart.x_axis.txPr = rich_text
    chart.x_axis.title.tx.rich.p[0].r[0].rPr = char_properties
    chart.y_axis.txPr = rich_text
    chart.y_axis.title.tx.rich.p[0].r[0].rPr = char_properties

    # 対数軸にする
    chart.x_axis.scaling.logBase = 10
    chart.y_axis.scaling.logBase = 10
    # 軸の位置は下端，左端に設定
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"

    # 補助目盛を追加
    chart.x_axis.minorGridlines = ChartLines()
    chart.y_axis.minorGridlines = ChartLines()

    # 凡例の位置とフォント
    chart.legend.legendPos = "tr"
    chart.legend.txPr = rich_text

    # ps曲線をシートに追加
    ws.add_chart(chart, 'B10')

def data_process(file_name,sheet_name,test_config, dim_srg):
    # データが始まる最初の行．排気性能測定シートのフォーマットが変わらなければ18のままでいい．
    start_row = 18
    # ブック読み込み

    # excel = win32com.client.Dispatch('Excel.Application')
    # excel.Visible = True
    # wb = excel.Workbooks.Open(file_name)
    # wb.Save()
    # excel.Quit()
    if file_name.split('.')[1] == 'xlsm':
        wb_tmp = xl.load_workbook(file_name, data_only=True, keep_vba=True)
        wb = xl.load_workbook(file_name, data_only=False, keep_vba=True)

    else:
        wb_tmp = xl.load_workbook(file_name, data_only=True)
        wb = xl.load_workbook(file_name, data_only=False)

    measured_data_sheet = wb_tmp[sheet_name]

    pspq, backpressure, list_pspq, dict_backpressure = data_sort(measured_data_sheet, start_row)
    wb_tmp.close()
    #データ整理とグラフを作るモジュールを呼び出す．
    if pspq == True:
        ws_pqps = wb.create_sheet('PQPS',1)
        write_pspq_data(ws_pqps, test_config, list_pspq, dim_srg)
        make_pspq_curve(ws_pqps, test_config)

    if backpressure == True:
        ws_backpressure = wb.create_sheet('Back pressure',1)
        write_backpressure_data(ws_backpressure, test_config, dict_backpressure,dim_srg)
        make_backpressure_curve(ws_backpressure, test_config, dict_backpressure)

    try:
        wb.save(file_name)
    except PermissionError:
        wb.save('new_book.xlsx')

        messagebox.showwarning('注意！','エクセルファイルが閉じられていないので，指定ファイルに保存できません．\n別ブックとして保存しました．')

    # wb = excel.Workbooks.Open(file_name)
    return pspq, backpressure
